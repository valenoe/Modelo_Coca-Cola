"""
lector_planificacion.py
Lector comun de los Excel de planificacion de la carpeta planning\.
Lo usan comparar_planificaciones.py y cargar_planificacion_real.py.

Estructura de cada hoja (una por auditor):

    fila 1      Planificacion mensual - <auditor>
    fila 2      Periodo | <Mes> | Tope mensual km | <tope>
    fila 3      Km planificados | <n> | Km disponibles | <n>
    fila 4      Dia | Fecha | Ruta | KLM | Clientes
    fila 5..N   un dia del mes cada una  <- RUTAS (tienen fecha en la col. B)
    'Total mensual'  / 'Km disponibles'  <- totales, se ignoran
    (sin fecha)  'Compra de Muestras...' <- EXTRAS: km que se hicieron pero
                                            sin dia asignado en la planificacion

Regla: una fila es ruta solo si la columna B trae una fecha del mes. Lo que
viene despues sin fecha se toma como extra, salvo las filas de total.
"""

import datetime

MESES = {"Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4, "Mayo": 5,
         "Junio": 6, "Julio": 7, "Agosto": 8, "Septiembre": 9,
         "Octubre": 10, "Noviembre": 11, "Diciembre": 12}
MES_NOMBRE = {v: k for k, v in MESES.items()}

# Filas que son totales, no rutas ni extras
ETIQUETAS_TOTAL = ("total", "km disponibles", "km planificados", "suma")


def norm(s):
    """Normaliza un nombre para comparar: sin espacios extra, en minusculas."""
    return " ".join(str(s).split()).lower()


def es_total(nombre):
    return norm(nombre).startswith(ETIQUETAS_TOTAL)


def dia_del_mes(valor, mes):
    """La columna B trae la fecha, como texto 'dd-mm-aaaa' o como fecha real.
    Retorna el dia si corresponde al mes pedido, o None si la fila no es un
    dia del calendario."""
    if isinstance(valor, (datetime.datetime, datetime.date)):
        return valor.day if valor.month == mes else None
    if isinstance(valor, str):
        partes = valor.strip().split("-")
        if len(partes) == 3 and all(p.isdigit() for p in partes):
            d, m, _a = (int(p) for p in partes)
            return d if m == mes else None
    return None


def leer_hoja(filas, mes):
    """Retorna (rutas, extras).

    rutas  : lista de (dia, nombre, km) — el dia permite que una misma ruta
             aparezca dos veces en el mes (rutas partidas en dos jornadas).
    extras : lista de (nombre, km) — km sin dia asignado.
    """
    rutas, extras = [], []
    for f in filas[4:]:
        if len(f) < 4:
            continue
        nombre, km = f[2], f[3]
        if not (isinstance(nombre, str) and nombre.strip()):
            continue
        if not isinstance(km, (int, float)):
            continue          # 'Feriado' y dias sin ruta vienen sin km
        nombre = nombre.strip()
        dia = dia_del_mes(f[1], mes)
        if dia is not None:
            rutas.append((dia, nombre, int(km)))
        elif not es_total(nombre):
            extras.append((nombre, int(km)))
    return rutas, extras


def leer_planificacion(path):
    """Lee un Excel de planificacion completo.

    Retorna (mes, plan, extras):
        mes    : numero de mes (1-12) o None si no se pudo determinar
        plan   : {auditor: [(dia, nombre_ruta, km), ...]}
        extras : {auditor: [(concepto, km), ...]}
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    mes, plan, extras = None, {}, {}

    for hoja in wb.sheetnames:
        filas = list(wb[hoja].iter_rows(min_row=1, values_only=True))
        if len(filas) < 5:
            continue
        if mes is None and len(filas[1]) > 1:
            mes = MESES.get(str(filas[1][1]).strip())
        if mes is None:
            continue

        r, e = leer_hoja(filas, mes)
        if r:
            plan[hoja] = r
        if e:
            extras[hoja] = e

    wb.close()
    return mes, plan, extras
