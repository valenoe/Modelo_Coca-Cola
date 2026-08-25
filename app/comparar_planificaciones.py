"""
comparar_planificaciones.py
Lee las planificaciones reales de la carpeta planning\ (junio, julio, agosto) y
las compara contra el catalogo de rutas de cada auditor en la BD.

Genera un Excel en planning\ con:
  - Una hoja por auditor: todas sus rutas + una columna por mes (1 si la ruta
    aparecio en la planificacion de ese mes, 0 si no). Si una ruta se uso mas
    de una vez en el mismo mes (rutas partidas en dos dias), aparece el numero
    de veces en vez de 1.
  - Una hoja "No calzaron" con todo lo que no cuadro: la ruta, el mes, el
    auditor, el km de la planificacion y el km de la BD.

Uso:
    python app\comparar_planificaciones.py
    python app\comparar_planificaciones.py <carpeta> <salida.xlsx>
"""

import sys
import os
import glob
import sqlite3
import collections
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAIZ     = os.path.dirname(BASE_DIR)
DB_PATH  = os.path.join(BASE_DIR, "roble.db")

sys.path.insert(0, BASE_DIR)
sys.stdout.reconfigure(encoding="utf-8")
from lector_planificacion import MESES, MES_NOMBRE, norm, leer_planificacion

VERDE_CAB  = "70AD47"
VERDE_CLAR = "E2EFDA"
ROJO_CLAR  = "FFC7CE"
AMARILLO   = "FFFF00"
GRIS       = "D9D9D9"
BLANCO     = "FFFFFF"


def cel(ws, f, c, v, bold=False, bg=None, color="000000", align="center",
        size=10, fmt=None):
    cc = ws.cell(row=f, column=c, value=v)
    cc.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cc.alignment = Alignment(horizontal=align, vertical="center")
    thin = Side(style="thin", color="BBBBBB")
    cc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if bg:
        cc.fill = PatternFill("solid", start_color=bg)
    if fmt:
        cc.number_format = fmt
    return cc


# ─────────────────────────────────────────────────────────────
# LECTURA
# ─────────────────────────────────────────────────────────────

def leer_rutas_bd():
    """{auditor: {nombre_norm: (id, numero, nombre, km)}}"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rutas = {}
    for r in conn.execute("""
        SELECT r.id, r.numero, r.nombre, r.distancia_km, a.nombre AS auditor
        FROM ruta r JOIN auditor a ON a.id = r.id_auditor
        ORDER BY a.nombre, r.numero
    """):
        rutas.setdefault(r["auditor"], {})[norm(r["nombre"])] = (
            r["id"], r["numero"], r["nombre"], r["distancia_km"])
    conn.close()
    return rutas


# ─────────────────────────────────────────────────────────────
# HOJAS DE SALIDA
# ─────────────────────────────────────────────────────────────

def hoja_auditor(wb, auditor, rutas_bd, planes, meses, extras):
    """Una hoja con las rutas del auditor y una columna 1/0 por mes."""
    partes = auditor.split()
    ws = wb.create_sheet(f"{partes[0]} {partes[1]}"[:31])

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 10
    for i, _ in enumerate(meses):
        ws.column_dimensions[chr(ord("D") + i)].width = 10
    ws.column_dimensions[chr(ord("D") + len(meses))].width = 12

    ncol_tot = 4 + len(meses)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol_tot)
    cel(ws, 1, 1, f"Uso de rutas por mes — {auditor}",
        bold=True, bg=VERDE_CAB, color=BLANCO, size=12)

    enc = ["N°", "Ruta", "Km BD"] + [MES_NOMBRE[m] for m in meses] + ["Total"]
    for j, t in enumerate(enc, start=1):
        cel(ws, 2, j, t, bold=True, bg=VERDE_CAB, color=BLANCO)

    # Conteo por mes: {mes: Counter(nombre_norm -> veces)}
    conteo = {}
    km_plan = {}
    for m in meses:
        c = collections.Counter()
        for _dia, nombre, km in planes.get(m, {}).get(auditor, []):
            c[norm(nombre)] += 1
            km_plan.setdefault(norm(nombre), {})[m] = km
        conteo[m] = c

    fila = 3
    catalogo = rutas_bd.get(auditor, {})
    vistos = set()

    # 1) Rutas del catalogo
    for clave, (_id, numero, nombre, km) in sorted(catalogo.items(), key=lambda kv: kv[1][1] or 0):
        vistos.add(clave)
        total = sum(conteo[m][clave] for m in meses)
        bg_fila = BLANCO if total else GRIS
        cel(ws, fila, 1, numero, bg=bg_fila)
        cel(ws, fila, 2, nombre, bg=bg_fila, align="left")
        cel(ws, fila, 3, km, bg=bg_fila, fmt="#,##0")
        for i, m in enumerate(meses):
            v = conteo[m][clave]
            bg = AMARILLO if v > 1 else (VERDE_CLAR if v else bg_fila)
            cel(ws, fila, 4 + i, v, bg=bg, bold=v > 1)
        cel(ws, fila, ncol_tot, total, bold=True,
            bg=VERDE_CLAR if total else GRIS)
        fila += 1

    # 2) Rutas que aparecen en las planificaciones pero NO estan en el catalogo
    faltantes = sorted(set(k for m in meses for k in conteo[m]) - vistos)
    if faltantes:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncol_tot)
        cel(ws, fila, 1, "NO ESTAN EN LA BD — se usaron en las planificaciones "
                         "pero no existen en el catalogo de rutas",
            bold=True, bg=ROJO_CLAR, align="left")
        fila += 1
        for clave in faltantes:
            nombre = next(n for m in meses for _d, n, _k in planes.get(m, {}).get(auditor, [])
                          if norm(n) == clave)
            kms = km_plan.get(clave, {})
            cel(ws, fila, 1, "", bg=ROJO_CLAR)
            cel(ws, fila, 2, nombre, bg=ROJO_CLAR, align="left")
            cel(ws, fila, 3, "", bg=ROJO_CLAR)
            for i, m in enumerate(meses):
                v = conteo[m][clave]
                cel(ws, fila, 4 + i, v, bg=ROJO_CLAR, bold=v > 1)
            cel(ws, fila, ncol_tot, sum(conteo[m][clave] for m in meses),
                bold=True, bg=ROJO_CLAR)
            fila += 1

    # 3) Extras: km sin dia asignado (no son rutas, pero suman al total del mes)
    conceptos = sorted({norm(n) for m in meses
                        for n, _ in extras.get(m, {}).get(auditor, [])})
    km_extras = {m: sum(km for _, km in extras.get(m, {}).get(auditor, []))
                 for m in meses}
    if conceptos:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=ncol_tot)
        cel(ws, fila, 1, "EXTRAS — km sin dia asignado, se suman al total del mes",
            bold=True, bg=AMARILLO, align="left")
        fila += 1
        for clave in conceptos:
            nombre = next(n for m in meses
                          for n, _ in extras.get(m, {}).get(auditor, [])
                          if norm(n) == clave)
            cel(ws, fila, 1, "", bg=AMARILLO)
            cel(ws, fila, 2, nombre, bg=AMARILLO, align="left")
            cel(ws, fila, 3, "", bg=AMARILLO)
            for i, m in enumerate(meses):
                km = sum(km for n, km in extras.get(m, {}).get(auditor, [])
                         if norm(n) == clave)
                cel(ws, fila, 4 + i, km or 0, bg=AMARILLO, fmt="#,##0")
            cel(ws, fila, ncol_tot, "", bg=AMARILLO)
            fila += 1

    # Totales
    cel(ws, fila, 1, "", bg=VERDE_CLAR)
    cel(ws, fila, 2, "Total dias asignados", bold=True, bg=VERDE_CLAR, align="left")
    cel(ws, fila, 3, "", bg=VERDE_CLAR)
    for i, m in enumerate(meses):
        cel(ws, fila, 4 + i, sum(conteo[m].values()), bold=True, bg=VERDE_CLAR)
    cel(ws, fila, ncol_tot, "", bg=VERDE_CLAR)

    fila += 1
    cel(ws, fila, 1, "", bg=VERDE_CLAR)
    cel(ws, fila, 2, "Total km (rutas + extras)", bold=True, bg=VERDE_CLAR, align="left")
    cel(ws, fila, 3, "", bg=VERDE_CLAR)
    for i, m in enumerate(meses):
        km_mes = sum(km for _d, _n, km in planes.get(m, {}).get(auditor, [])) + km_extras[m]
        cel(ws, fila, 4 + i, km_mes, bold=True, bg=VERDE_CLAR, fmt="#,##0")
    cel(ws, fila, ncol_tot, "", bg=VERDE_CLAR)


def hoja_no_calzaron(wb, filas_error):
    ws = wb.create_sheet("No calzaron", 0)
    anchos = {"A": 20, "B": 12, "C": 44, "D": 16, "E": 12, "F": 34}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:F1")
    cel(ws, 1, 1, "Rutas que no calzaron con la base de datos",
        bold=True, bg=VERDE_CAB, color=BLANCO, size=12)

    enc = ["Auditor", "Mes", "Ruta", "Km planificacion", "Km BD", "Motivo"]
    for j, t in enumerate(enc, start=1):
        cel(ws, 2, j, t, bold=True, bg=VERDE_CAB, color=BLANCO)

    fila = 3
    for e in filas_error:
        bg = ROJO_CLAR if e["motivo"].startswith("No existe") else AMARILLO
        cel(ws, fila, 1, e["auditor"],  bg=bg, align="left")
        cel(ws, fila, 2, e["mes"],      bg=bg)
        cel(ws, fila, 3, e["ruta"],     bg=bg, align="left")
        cel(ws, fila, 4, e["km_plan"],  bg=bg, fmt="#,##0")
        cel(ws, fila, 5, e["km_bd"] if e["km_bd"] is not None else "—",
            bg=bg, fmt="#,##0")
        cel(ws, fila, 6, e["motivo"],   bg=bg, align="left")
        fila += 1

    if not filas_error:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
        cel(ws, 3, 1, "Todo calzo — ninguna diferencia encontrada",
            bold=True, bg=VERDE_CLAR)


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    carpeta = sys.argv[1] if len(sys.argv) > 1 else os.path.join(RAIZ, "planning")
    salida  = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        carpeta, "Comparacion_Rutas.xlsx")

    # Se omiten los "Comparacion*" (los genera este mismo script) y los "~$"
    # (temporales que crea Excel al tener un archivo abierto).
    archivos = [p for p in sorted(glob.glob(os.path.join(carpeta, "*.xlsx")))
                if not os.path.basename(p).startswith(("~$", "Comparacion"))]
    if not archivos:
        print(f"No se encontraron planificaciones en {carpeta}")
        sys.exit(1)

    rutas_bd = leer_rutas_bd()

    planes  = {}    # {mes: {auditor: [(ruta, km)]}}
    extras  = {}    # {mes: {auditor: [(concepto, km)]}}  km sin dia asignado
    print("Leyendo planificaciones:")
    for p in archivos:
        mes, plan, extra = leer_planificacion(p)
        if mes is None:
            print(f"  ⚠️  {os.path.basename(p)}: no se pudo determinar el mes, se omite")
            continue
        if mes in planes:
            print(f"  ⚠️  {os.path.basename(p)}: {MES_NOMBRE[mes]} ya estaba cargado, se omite")
            continue
        planes[mes] = plan
        extras[mes] = extra
        dias = sum(len(v) for v in plan.values())
        print(f"  {os.path.basename(p):<45} {MES_NOMBRE[mes]:<10} "
              f"{len(plan)} auditores, {dias} dias asignados")

    meses = sorted(planes)

    # Detectar lo que no calza
    errores = []
    for m in meses:
        for auditor, asignaciones in planes[m].items():
            catalogo = rutas_bd.get(auditor, {})
            reportados = set()
            for _dia, nombre, km in asignaciones:
                clave = norm(nombre)
                if clave in reportados:
                    continue
                if clave not in catalogo:
                    reportados.add(clave)
                    errores.append({
                        "auditor": auditor, "mes": MES_NOMBRE[m], "ruta": nombre,
                        "km_plan": km, "km_bd": None,
                        "motivo": "No existe en la BD"})
                elif catalogo[clave][3] != km:
                    reportados.add(clave)
                    errores.append({
                        "auditor": auditor, "mes": MES_NOMBRE[m], "ruta": nombre,
                        "km_plan": km, "km_bd": catalogo[clave][3],
                        "motivo": f"Km distinto (diferencia {km - catalogo[clave][3]:+d})"})

    # Auditores: los de la BD que aparecen en alguna planificacion
    auditores = sorted({a for m in meses for a in planes[m]})

    wb = Workbook()
    wb.remove(wb.active)
    for auditor in auditores:
        hoja_auditor(wb, auditor, rutas_bd, planes, meses, extras)
    hoja_no_calzaron(wb, errores)

    wb.save(salida)

    print(f"\nGuardado: {salida}")
    print(f"  Meses: {', '.join(MES_NOMBRE[m] for m in meses)}")
    print(f"  Hojas: {len(auditores)} auditores + 'No calzaron'")
    print(f"  Diferencias detectadas: {len(errores)}")
    for e in errores:
        print(f"    {e['auditor']:<20} {e['mes']:<8} {e['ruta']:<40} "
              f"plan={e['km_plan']:>4}  bd={e['km_bd'] if e['km_bd'] is not None else '—':>4}  {e['motivo']}")


if __name__ == "__main__":
    main()
