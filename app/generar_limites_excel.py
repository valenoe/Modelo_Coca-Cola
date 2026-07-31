"""
generar_limites_excel.py
Genera un Excel con los cálculos de la Sección 7 para un mes/año y precio de
bencina dados: km estimado, km con bono (tope), bono por bencina y pago km.
Útil para revisar cómo cambian los bonos cuando cambia el precio de la bencina.

Uso:
    python app\\generar_limites_excel.py <precio_bencina> <mes> <anio> [salida.xlsx]

Ejemplo julio 2026 con bencina $1.496:
    python app\\generar_limites_excel.py 1496 7 2026 Limites_Julio_2026.xlsx
"""

import sys
import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
from motor import calcular_km_con_bono

DB_PATH = os.path.join(BASE_DIR, "roble.db")

MESES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
         7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre",
         11: "Noviembre", 12: "Diciembre"}

VERDE_CAB  = "70AD47"
VERDE_CLAR = "E2EFDA"
GRIS       = "D9D9D9"
AMARILLO   = "FFFF00"
BLANCO     = "FFFFFF"


def cel(ws, fila, col, valor, bold=False, bg=None, color="000000",
        align="center", size=10, num_fmt=None):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font      = Font(name="Arial", bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    thin        = Side(style="thin", color="AAAAAA")
    c.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if num_fmt:
        c.number_format = num_fmt
    return c


def main():
    if len(sys.argv) < 4:
        print(r"Uso: python app\generar_limites_excel.py <precio_bencina> <mes> <anio> [salida.xlsx]")
        sys.exit(1)

    precio = float(sys.argv[1])
    mes_arg = sys.argv[2]
    if mes_arg.isdigit():
        mes = int(mes_arg)
    else:
        nombre2num = {v.lower(): k for k, v in MESES.items()}
        mes = nombre2num.get(mes_arg.strip().lower())
        if mes is None:
            print(f"Mes no reconocido: {mes_arg!r}. Usa el número (1-12) o el nombre (ej. Junio).")
            sys.exit(1)
    anio   = int(sys.argv[3])
    salida = sys.argv[4] if len(sys.argv) > 4 else f"Limites_{MESES[mes]}_{anio}.xlsx"
    if not salida.endswith(".xlsx"):
        salida += ".xlsx"

    # Contexto: bencina anterior, tarifa y presupuesto del mes
    conn = sqlite3.connect(DB_PATH)
    cfg  = {k: float(v) for k, v in conn.execute("SELECT clave, valor FROM config")}
    presu = conn.execute(
        "SELECT presupuesto_total FROM presupuesto_anual WHERE anio=? AND mes=?",
        (anio, mes)).fetchone()
    conn.close()
    presupuesto = presu[0] if presu else 0

    resultados, M21, M15, M14_act = calcular_km_con_bono(anio, mes, precio)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESES[mes]} {anio}"

    anchos = {"A": 24, "B": 14, "C": 14, "D": 13, "E": 14, "F": 14}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # Título
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 24
    cel(ws, 1, 1, f"Cálculo de Km y Bono — {MESES[mes]} {anio}",
        bold=True, bg=VERDE_CAB, color=BLANCO, size=13)

    # Contexto (bencina / presupuesto / tarifa)
    ws.merge_cells("A2:C2"); ws.merge_cells("D2:F2")
    cel(ws, 2, 1, f"Bencina actual: ${precio:,.0f}/litro", bold=True, bg=VERDE_CLAR, align="left")
    cel(ws, 2, 4, f"Bencina anterior: ${cfg.get('precio_bencina_anterior', 0):,.0f}/litro", bg=VERDE_CLAR, align="left")
    ws.merge_cells("A3:C3"); ws.merge_cells("D3:F3")
    cel(ws, 3, 1, f"Presupuesto del mes: ${presupuesto:,.0f}", bold=True, bg=VERDE_CLAR, align="left")
    cel(ws, 3, 4, f"Tarifa: {cfg.get('tarifa_actual', 0):,.0f} (ant. {cfg.get('tarifa_anterior', 0):,.0f})", bg=VERDE_CLAR, align="left")

    # Encabezados de la tabla
    enc = ["Auditor", "Km estimado", "Km con bono", "Diferencia km", "Bono ($)", "Pago km ($)"]
    for j, t in enumerate(enc, start=1):
        cel(ws, 5, j, t, bold=True, bg=VERDE_CAB, color=BLANCO)

    fila = 6
    tot_km_bono = tot_bono = tot_pago = 0
    # activos primero, luego inactivos
    items = sorted(resultados.items(), key=lambda kv: (0 if kv[1]["activo"] else 1, kv[1]["nombre"]))
    for id_a, v in items:
        if not v:
            continue
        bg = BLANCO if v["activo"] else GRIS
        nombre = v["nombre"] + ("" if v["activo"] else " (inactivo)")
        cel(ws, fila, 1, nombre, bg=bg, align="left")
        cel(ws, fila, 2, round(v["G14_act"]),     bg=bg, num_fmt="#,##0")
        cel(ws, fila, 3, round(v["km_con_bono"]), bg=AMARILLO if v["activo"] else bg, num_fmt="#,##0", bold=v["activo"])
        cel(ws, fila, 4, round(v["diferencia_km"]), bg=bg, num_fmt="#,##0")
        cel(ws, fila, 5, round(v["bono"]),        bg=bg, num_fmt="#,##0")
        cel(ws, fila, 6, round(v["pago_km"]),     bg=bg, num_fmt="#,##0")
        if v["activo"]:
            tot_km_bono += v["km_con_bono"]
            tot_bono    += v["bono"]
            tot_pago    += v["pago_km"]
        fila += 1

    # Totales (solo activos)
    cel(ws, fila, 1, "Total activos", bold=True, bg=VERDE_CLAR, align="left")
    cel(ws, fila, 2, "", bg=VERDE_CLAR)
    cel(ws, fila, 3, round(tot_km_bono), bold=True, bg=VERDE_CLAR, num_fmt="#,##0")
    cel(ws, fila, 4, "", bg=VERDE_CLAR)
    cel(ws, fila, 5, round(tot_bono), bold=True, bg=VERDE_CLAR, num_fmt="#,##0")
    cel(ws, fila, 6, round(tot_pago), bold=True, bg=VERDE_CLAR, num_fmt="#,##0")

    wb.save(salida)
    print(f"Guardado: {salida}")
    print(f"  Bencina ${precio:,.0f} | {MESES[mes]} {anio} | {len([1 for _,v in items if v and v['activo']])} auditores activos")


if __name__ == "__main__":
    main()
