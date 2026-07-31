"""
cargar_mayo.py
Carga las rutas de mayo 2026 desde el Excel de planificación
como programación inicial en la BD.

Ejecutar UNA SOLA VEZ antes de empezar a usar el sistema.

Uso:
    python app\cargar_mayo.py
"""

import sqlite3
import os
from openpyxl import load_workbook

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(BASE_DIR, "roble.db")
EXCEL_MAYO = os.path.join(BASE_DIR, "28052026_PDR_Mayo.xlsx")

ANIO = 2026
MES  = 5

def main():
    if not os.path.exists(EXCEL_MAYO):
        print(f"❌ No se encontró el archivo: {EXCEL_MAYO}")
        print("   Ponlo en la carpeta app\\")
        return

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    wb = load_workbook(EXCEL_MAYO, read_only=True)
    hojas_auditores = [s for s in wb.sheetnames if s != "Resumen"]

    total_cargadas = 0
    total_no_encontradas = 0

    for sheet in hojas_auditores:
        # Buscar auditor en BD
        aud = c.execute("SELECT id FROM auditor WHERE nombre=?", (sheet,)).fetchone()
        if not aud:
            print(f"⚠️  Auditor '{sheet}' no encontrado en la BD, se omite")
            continue
        id_auditor = aud["id"]

        ws = wb[sheet]
        rutas_cargadas      = []
        rutas_no_encontradas = []

        for fila in ws.iter_rows(min_row=5, values_only=True):
            # Columna C = nombre ruta, columna D = km
            nombre_ruta = fila[2] if len(fila) > 2 else None
            if not nombre_ruta or not isinstance(nombre_ruta, str):
                continue
            nombre_ruta = nombre_ruta.strip()

            # Buscar la ruta en la BD por nombre exacto
            ruta = c.execute("""
                SELECT id FROM ruta
                WHERE id_auditor=? AND nombre=?
            """, (id_auditor, nombre_ruta)).fetchone()

            if not ruta:
                # Buscar por nombre similar (ignorando mayúsculas/espacios extra)
                ruta = c.execute("""
                    SELECT id FROM ruta
                    WHERE id_auditor=? AND LOWER(TRIM(nombre))=LOWER(TRIM(?))
                """, (id_auditor, nombre_ruta)).fetchone()

            if not ruta:
                rutas_no_encontradas.append(nombre_ruta)
                continue

            # Registrar como programada en mayo 2026
            c.execute("""
                INSERT OR IGNORE INTO programacion_mensual (id_auditor, anio, mes, id_ruta)
                VALUES (?, ?, ?, ?)
            """, (id_auditor, ANIO, MES, ruta["id"]))
            rutas_cargadas.append(nombre_ruta)
            total_cargadas += 1

        print(f"\n{sheet}: {len(rutas_cargadas)} rutas cargadas")
        if rutas_no_encontradas:
            print(f"  ⚠️  No encontradas en BD ({len(rutas_no_encontradas)}):")
            for r in rutas_no_encontradas:
                print(f"     - {r}")
            total_no_encontradas += len(rutas_no_encontradas)

    conn.commit()
    wb.close()
    conn.close()

    print(f"\n{'='*50}")
    print(f"✅ Mayo 2026 cargado: {total_cargadas} rutas")
    if total_no_encontradas:
        print(f"⚠️  {total_no_encontradas} rutas no encontradas — revisar nombres")
    print(f"{'='*50}\n")

if __name__ == "__main__":
    main()