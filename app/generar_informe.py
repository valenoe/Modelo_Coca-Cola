"""
generar_informe.py
Lee el txt generado por probar_semestre.py y produce un Excel
con una hoja por auditor en formato de planificacion mensual.

Uso:
    python app\generar_informe.py junio7.txt
    python app\generar_informe.py junio7.txt informe_junio.xlsx
"""

import sys
import re
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CLIENTES_MIN = 15

# Feriados del mes — agregar fechas como datetime.date(anio, mes, dia)
# El script los saltará igual que sábados y domingos
import datetime
FERIADOS = set()   # Agosto 2026: 15-ago (Asunción) cae sábado, no descuenta día hábil

VERDE_CAB  = "70AD47"
VERDE_CLAR = "E2EFDA"
GRIS_FIN   = "D9D9D9"
AMARILLO   = "FFFF00"
BLANCO     = "FFFFFF"

DIAS_ES = {0:"Lunes", 1:"Martes", 2:"Miercoles", 3:"Jueves",
           4:"Viernes", 5:"Sabado", 6:"Domingo"}

MESES = {"Enero":1,"Febrero":2,"Marzo":3,"Abril":4,"Mayo":5,"Junio":6,
         "Julio":7,"Agosto":8,"Septiembre":9,"Octubre":10,
         "Noviembre":11,"Diciembre":12}


def parsear_txt(path):
    try:
        with open(path, encoding="utf-16") as f:
            lines = f.readlines()
    except UnicodeError:
        with open(path, encoding="utf-8") as f:
            lines = f.readlines()

    auditores   = []
    auditor     = None
    anio_global = 2026

    for l in lines:
        m = re.search(r'(\d{4})', l)
        if m:
            anio_global = int(m.group(1))
            break

    i = 0
    while i < len(lines):
        l       = lines[i].rstrip("\n")
        stripped = l.strip()

        # Separador --- => siguiente línea es nombre de auditor
        if re.match(r'^-{20,}$', stripped) and i + 1 < len(lines):
            nombre_cand = lines[i + 1].strip()
            if (re.match(r'^[A-Z][a-zA-Z]+ [A-Z]', nombre_cand)
                    and not re.search(r'[\d|=]', nombre_cand)
                    and "Auditor" not in nombre_cand):
                if auditor is not None:
                    auditores.append(auditor)
                auditor = {"nombre": nombre_cand, "rutas": []}

        # Resumen del mes
        m = re.match(
            r'(\w+)\s+--\s+(\d+)\s+dias.*?Km bono:\s*([\d,]+)'
            r'.*?Km rutas:\s*([\d,]+).*?Saldo:\s*([+-]?[\d,]+)',
            stripped)
        if m and auditor is not None:
            auditor["mes_nombre"] = m.group(1)
            auditor["anio"]       = anio_global
            auditor["mes_num"]    = MESES.get(m.group(1), 6)
            auditor["dias"]       = int(m.group(2))
            auditor["km_bono"]    = int(m.group(3).replace(",", ""))
            auditor["km_rutas"]   = int(m.group(4).replace(",", ""))
            auditor["saldo"]      = int(m.group(5).replace(",", ""))

        # Ruta
        m = re.match(r'\s{4,}(.+?)\s{2,}(\d+)\s+km\s*$', l)
        if m and auditor is not None and "mes_num" in auditor:
            auditor["rutas"].append((m.group(1).strip(), int(m.group(2))))

        i += 1

    if auditor is not None and "mes_num" in auditor:
        auditores.append(auditor)

    # Filtrar solo los que tienen rutas reales
    return [a for a in auditores if a.get("rutas")]


def cel(ws, fila, col, valor, bold=False, bg=None, color="000000",
        align="center", size=10, num_fmt=None):
    cell           = ws.cell(row=fila, column=col, value=valor)
    cell.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    thin           = Side(style="thin", color="AAAAAA")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def crear_hoja(wb, aud):
    nombre   = aud["nombre"]
    anio     = aud["anio"]
    mes_num  = aud["mes_num"]
    mes_nom  = aud["mes_nombre"]
    km_bono  = aud["km_bono"]
    rutas    = aud["rutas"]

    partes      = nombre.split()
    nombre_hoja = f"{partes[0]} {partes[1]}"[:31]
    ws          = wb.create_sheet(nombre_hoja)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16

    # Fila 1 — titulo
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 22
    cel(ws, 1, 1, f"Planificacion mensual - {nombre}",
        bold=True, bg=VERDE_CAB, color=BLANCO, size=12)

    # Fila 2 — periodo / tope
    ws.row_dimensions[2].height = 18
    cel(ws, 2, 1, "Periodo",         bold=True, bg=VERDE_CLAR)
    cel(ws, 2, 2, mes_nom,           bg=BLANCO)
    cel(ws, 2, 3, "Tope mensual km", bold=True, bg=VERDE_CLAR)
    cel(ws, 2, 4, km_bono,           bg=BLANCO, num_fmt="#,##0")
    cel(ws, 2, 5, "",                bg=BLANCO)

    # Fila 3 — km planificados / disponibles
    ws.row_dimensions[3].height = 18
    cel(ws, 3, 1, "Km planificados", bold=True, bg=VERDE_CLAR)
    import calendar
    dias_mes     = calendar.monthrange(anio, mes_num)[1]
    ultima_fila  = 4 + dias_mes   # fila 5 = día 1, fila 5+dias-1 = último día
    cel(ws, 3, 2, f"=SUM(D5:D{ultima_fila})", bg=BLANCO, num_fmt="#,##0")
    cel(ws, 3, 3, "Km disponibles",  bold=True, bg=VERDE_CLAR)
    cel(ws, 3, 4, f"=D2-B3",         bg=AMARILLO, num_fmt="#,##0", color="FF0000")
    cel(ws, 3, 5, "",                bg=BLANCO)

    # Fila 4 — encabezados
    ws.row_dimensions[4].height = 18
    cel(ws, 4, 1, "Dia",                 bold=True, bg=VERDE_CAB, color=BLANCO)
    cel(ws, 4, 2, "Fecha",               bold=True, bg=VERDE_CAB, color=BLANCO)
    cel(ws, 4, 3, "Ruta",                bold=True, bg=VERDE_CAB, color=BLANCO)
    cel(ws, 4, 4, "KLM",                 bold=True, bg=VERDE_CAB, color=BLANCO)
    cel(ws, 4, 5, "Clientes (Minimos)",  bold=True, bg=VERDE_CAB, color=BLANCO)

    # Asignar rutas a dias habiles en orden (excluye feriados)
    dias_habiles = []
    fecha = datetime.date(anio, mes_num, 1)
    while fecha.month == mes_num:
        if fecha.weekday() < 5 and fecha not in FERIADOS:
            dias_habiles.append(fecha)
        fecha += datetime.timedelta(days=1)

    rutas_por_fecha = {}
    for idx, ruta in enumerate(rutas):
        if idx < len(dias_habiles):
            rutas_por_fecha[dias_habiles[idx]] = ruta

    # Dias del mes
    fila  = 5
    fecha = datetime.date(anio, mes_num, 1)
    while fecha.month == mes_num:
        ws.row_dimensions[fila].height = 16
        es_feriado = fecha in FERIADOS
        es_fin = fecha.weekday() >= 5 or es_feriado
        bg     = GRIS_FIN if es_fin else BLANCO
        dow    = DIAS_ES[fecha.weekday()]
        ruta   = rutas_por_fecha.get(fecha)

        cel(ws, fila, 1, dow,                          bg=bg)
        cel(ws, fila, 2, fecha.strftime("%d-%m-%Y"),    bg=bg)
        cel(ws, fila, 3, ruta[0] if ruta else "",       bg=bg, align="left")
        cel(ws, fila, 4, ruta[1] if ruta else "",       bg=bg, num_fmt="#,##0")
        cel(ws, fila, 5, CLIENTES_MIN if ruta else "",  bg=bg)
        if es_feriado and not ruta:
            cel(ws, fila, 3, "Feriado", bg=GRIS_FIN, color="888888", align="center")

        fila  += 1
        fecha += datetime.timedelta(days=1)

    # Total
    ws.row_dimensions[fila].height = 18
    cel(ws, fila, 1, "Total mensual", bold=True, bg=VERDE_CLAR, align="left")
    cel(ws, fila, 2, "",              bg=VERDE_CLAR)
    cel(ws, fila, 3, "",              bg=VERDE_CLAR)
    cel(ws, fila, 4, f"=SUM(D5:D{fila-1})", bold=True, bg=VERDE_CLAR, num_fmt="#,##0")
    cel(ws, fila, 5, f"=SUM(E5:E{fila-1})", bold=True, bg=VERDE_CLAR, num_fmt="#,##0")

    fila += 1
    ws.row_dimensions[fila].height = 18
    cel(ws, fila, 1, "Km disponibles", bold=True, bg=AMARILLO, align="left")
    cel(ws, fila, 2, "",               bg=AMARILLO)
    cel(ws, fila, 3, "",               bg=AMARILLO)
    cel(ws, fila, 4, f"=D2-D{fila-1}", bold=True, bg=AMARILLO, num_fmt="#,##0", color="FF0000")
    cel(ws, fila, 5, "",               bg=AMARILLO)


def main():
    if len(sys.argv) < 2:
        print("Uso: python app\\generar_informe.py <resultado.txt> [salida.xlsx]")
        sys.exit(1)

    txt_path = sys.argv[1]
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else txt_path.replace(".txt", ".xlsx")
    if not xlsx_out.endswith(".xlsx"):
        xlsx_out += ".xlsx"

    print(f"Leyendo {txt_path}...")
    auditores = parsear_txt(txt_path)
    print(f"  {len(auditores)} auditores: {[a['nombre'] for a in auditores]}")

    wb = Workbook()
    wb.remove(wb.active)

    for aud in auditores:
        print(f"  {aud['nombre']}: {len(aud['rutas'])} rutas | "
              f"km_bono={aud['km_bono']} | saldo={aud['saldo']}")
        crear_hoja(wb, aud)

    wb.save(xlsx_out)
    print(f"Guardado: {xlsx_out}")


if __name__ == "__main__":
    main()