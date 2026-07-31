"""
generar_paso_a_paso.py
Genera un Excel con el PASO A PASO del cálculo de la Sección 7 (km con bono),
con los auditores en columnas y cada paso intermedio en una fila — igual que el
Excel de referencia. Sirve para comparar fila por fila y ubicar diferencias.

Uso:
    python app\\generar_paso_a_paso.py <precio_bencina> <mes> <anio> [salida.xlsx]

Ejemplo julio 2026, bencina $1.496:
    python app\\generar_paso_a_paso.py 1496 7 2026 PasoAPaso_Julio_2026.xlsx
"""

import sys
import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE_DIR)
from motor import calcular_km_con_bono

DB_PATH = os.path.join(BASE_DIR, "roble.db")

MESES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
         7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre",
         11: "Noviembre", 12: "Diciembre"}

# Orden de columnas (igual que el Excel de referencia / histórico)
ORDEN = ["Carlos Acevedo", "Marco Contreras", "Samuel Inostroza",
         "Cristian Lizama", "Pablo Villarroel", "Mauricio Picart"]

AZUL_OSC = "1F4E79"
AZUL_CLA = "2E75B6"
CELESTE  = "DDEBF7"
AMARILLO = "FFFF00"
VERDE    = "E2EFDA"
GRIS     = "D9D9D9"
BLANCO   = "FFFFFF"


def cel(ws, f, c, v, bold=False, bg=None, color="000000", align="center",
        size=10, fmt=None):
    cc = ws.cell(row=f, column=c, value=v)
    cc.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cc.alignment = Alignment(horizontal=align, vertical="center")
    thin = Side(style="thin", color="BBBBBB")
    cc.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if bg:
        cc.fill = PatternFill("solid", start_color=bg)
    if fmt:
        cc.number_format = fmt
    return cc


def main():
    if len(sys.argv) < 4:
        print(r"Uso: python app\generar_paso_a_paso.py <precio_bencina> <mes> <anio> [salida.xlsx]")
        sys.exit(1)

    precio  = float(sys.argv[1])
    mes_arg = sys.argv[2]
    if mes_arg.isdigit():
        mes = int(mes_arg)
    else:
        n2n = {v.lower(): k for k, v in MESES.items()}
        mes = n2n.get(mes_arg.strip().lower())
        if mes is None:
            print(f"Mes no reconocido: {mes_arg!r}")
            sys.exit(1)
    anio   = int(sys.argv[3])
    salida = sys.argv[4] if len(sys.argv) > 4 else f"PasoAPaso_{MESES[mes]}_{anio}.xlsx"
    if not salida.endswith(".xlsx"):
        salida += ".xlsx"

    # Constantes y agregados
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cfg = {r["clave"]: float(r["valor"]) for r in conn.execute("SELECT clave, valor FROM config")}
    M20_ant = conn.execute("SELECT presupuesto_total FROM presupuesto_anual WHERE anio=? AND mes=?", (anio - 1, mes)).fetchone()[0]
    M20_act = conn.execute("SELECT presupuesto_total FROM presupuesto_anual WHERE anio=? AND mes=?", (anio, mes)).fetchone()[0]
    ciudades = {r["nombre"]: r["ciudad"] for r in conn.execute("SELECT nombre, ciudad FROM auditor")}
    ids      = {r["nombre"]: r["id"]     for r in conn.execute("SELECT nombre, id FROM auditor")}
    conn.close()

    T     = cfg["tarifa_actual"]
    T_ant = cfg["tarifa_anterior"]
    P_ant = cfg["precio_bencina_anterior"]

    res, M21, M15, M14_act = calcular_km_con_bono(anio, mes, precio)
    M14_ant = sum(v["G14_ant"] for v in res.values() if v)
    factor  = M21 / (T * M15)

    # Auditores presentes, en el orden de referencia
    cols = [n for n in ORDEN if n in ids and res.get(ids[n])]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESES[mes]} {anio}"

    ws.column_dimensions["A"].width = 30
    for i in range(len(cols)):
        ws.column_dimensions[chr(ord("B") + i)].width = 15
    ws.column_dimensions[chr(ord("B") + len(cols))].width = 15

    ncol_tot = 2 + len(cols)  # columna de Total

    # ── Bloque de constantes ──────────────────────────────────
    f = 1
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=ncol_tot)
    cel(ws, f, 1, f"Paso a paso — Cálculo de Km con Bono — {MESES[mes]} {anio}",
        bold=True, bg=AZUL_OSC, color=BLANCO, size=13)
    f += 1

    consts = [
        ("Tarifa actual / anterior", f"{T:,.0f} / {T_ant:,.0f}"),
        ("Bencina actual / anterior / diferencia", f"${precio:,.0f} / ${P_ant:,.0f} / ${precio - P_ant:,.0f}"),
        (f"Presupuesto {anio-1} / {anio} (M20)", f"${M20_ant:,.0f} / ${M20_act:,.0f}"),
        ("Suma km 2025 (M14_ant) / proyectado (M14_act)", f"{M14_ant:,.0f} / {M14_act:,.0f}"),
        ("Suma Valor 140 (M15) / Suma Pago sin bono (M21)", f"{M15:,.0f} / {M21:,.0f}"),
        ("Factor Km con bono = M21 / (Tarifa × M15)", f"{factor:.5f}"),
        ("Fórmula: Km con bono = Factor × Valor 140", "(de cada auditor)"),
    ]
    for etiqueta, valor in consts:
        ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=2)
        cel(ws, f, 1, etiqueta, bold=True, bg=VERDE, align="left")
        ws.merge_cells(start_row=f, start_column=3, end_row=f, end_column=ncol_tot)
        cel(ws, f, 3, valor, bg=BLANCO, align="left")
        f += 1

    f += 1  # fila en blanco

    # ── Encabezado de la tabla ────────────────────────────────
    ws.merge_cells(start_row=f, start_column=2, end_row=f, end_column=1 + len(cols))
    cel(ws, f, 2, "Auditores", bold=True, bg=AZUL_OSC, color=BLANCO, size=12)
    f += 1
    fila_nom = f
    cel(ws, f, 1, "", bg=AZUL_CLA)
    for j, n in enumerate(cols):
        et = n + ("" if res[ids[n]]["activo"] else " (inact.)")
        cel(ws, f, 2 + j, et, bold=True, bg=AZUL_CLA, color=BLANCO)
    cel(ws, f, ncol_tot, "Total", bold=True, bg=AZUL_OSC, color=BLANCO)
    f += 1
    # Ciudad
    cel(ws, f, 1, "Ciudad", bold=True, bg=CELESTE, align="left")
    for j, n in enumerate(cols):
        cel(ws, f, 2 + j, ciudades.get(n, ""), bg=BLANCO)
    cel(ws, f, ncol_tot, "", bg=CELESTE)
    f += 1

    # ── Filas de pasos ────────────────────────────────────────
    # (etiqueta, clave en res, formato, suma_total, resaltar)
    pasos = [
        (f"Km {MESES[mes]} 2025  (G14_ant)",       "G14_ant",       "#,##0",   True,  False),
        (f"Km proyectado {anio}  (G14_act)",        "G14_act",       "#,##0",   True,  False),
        ("Valor 140 — km a tarifa 140  (G15)",      "G15",           "#,##0",   True,  False),
        ("Rendimiento (litros)  (G18)",             "G18",           "#,##0",   True,  False),
        ("Bono estabilización ($)  (G19)",          "G19",           "$#,##0",  True,  False),
        ("Pago 120 × km ($)  (G20)",                "G20",           "$#,##0",  True,  False),
        ("Pago 140 sin bono ($)  (G21)",            "G21",           "$#,##0",  True,  False),
        ("KM CON BONO  (tope)",                     "km_con_bono",   "#,##0",   True,  True),
        ("Diferencia de km  (proy − tope)",         "diferencia_km", "#,##0",   True,  False),
    ]
    for etiqueta, clave, fmt, suma, resaltar in pasos:
        bg_lbl = AMARILLO if resaltar else CELESTE
        cel(ws, f, 1, etiqueta, bold=resaltar, bg=bg_lbl, align="left")
        total = 0.0
        for j, n in enumerate(cols):
            v = res[ids[n]][clave]
            bg = AMARILLO if resaltar else (BLANCO if res[ids[n]]["activo"] else GRIS)
            cel(ws, f, 2 + j, round(v), bold=resaltar, bg=bg, fmt=fmt)
            total += v
        if suma:
            cel(ws, f, ncol_tot, round(total), bold=True, bg=(AMARILLO if resaltar else VERDE), fmt=fmt)
        else:
            cel(ws, f, ncol_tot, "", bg=VERDE)
        f += 1

    wb.save(salida)
    print(f"Guardado: {salida}")
    print(f"  {MESES[mes]} {anio} | Bencina ${precio:,.0f} | {len(cols)} auditores")
    print(f"  Factor Km con bono = M21/(T*M15) = {factor:.5f}")


if __name__ == "__main__":
    main()
