"""
inicializar_bd.py
Crea y carga la base de datos SQLite desde los archivos Excel.
Ejecutar UNA SOLA VEZ, o cuando cambien las rutas de un auditor.

Uso:
    python app/inicializar_bd.py
"""

import sqlite3
import os
from openpyxl import load_workbook

# ─────────────────────────────────────────
# RUTAS DE ARCHIVOS
# ─────────────────────────────────────────

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DB_PATH   = os.path.join(BASE_DIR, "roble.db")

# Pon los Excel en la misma carpeta app\
EXCEL_RUTAS    = os.path.join(BASE_DIR, "Rutas___Kilometros_Auditores.xlsx")
EXCEL_HISTORICO = os.path.join(BASE_DIR, "_ABR26_Roble_Ajuste_KM_macros.xlsm")

# ─────────────────────────────────────────
# DATOS FIJOS
# ─────────────────────────────────────────

AUDITORES_ACTIVOS = {
    "Carlos Acevedo":   {"rendimiento": 11, "ciudad": "Talca"},
    "Samuel Inostroza": {"rendimiento": 17, "ciudad": "Curicó"},
    "Cristian Lizama":  {"rendimiento": 17, "ciudad": "San Fernando"},
    "Pablo Villarroel": {"rendimiento": 12, "ciudad": "Talca"},
    "Mauricio Picart":  {"rendimiento": 15, "ciudad": "Talca"},
}

# Marco Contreras: inactivo, pero su histórico se necesita para los cálculos
AUDITORES_INACTIVOS = {
    "Marco Contreras": {"rendimiento": 15, "ciudad": "Parral"},
}

# Orden de columnas en el Excel histórico
AUDITORES_ORDEN_HIST = [
    "Carlos Acevedo",
    "Marco Contreras",
    "Samuel Inostroza",
    "Cristian Lizama",
    "Pablo Villarroel",
    "Mauricio Picart",
]

MESES_NUM = {
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
    "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8,
    "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
}

# Presupuesto mensual total (M20) por año
PRESUPUESTO = {
    2025: {
        1: 2162160, 2: 951960,  3: 1402920, 4: 2101320,
        5: 1674600, 6: 2112840, 7: 1811640, 8: 1575480,
        9: 1401000, 10: 2138160, 11: 1535880, 12: 1804560,
    },
    2026: {
        1: 1829760, 2: 1200360, 3: 1316640, 4: 1662720,
        5: 1978340, 6: 1812100, 7: 1812100, 8: 1812100,
        9: 1812100, 10: 1812100, 11: 1812100, 12: 1812100,
    },
}

# Parámetros configurables
CONFIG = {
    "tarifa_actual":            "140",
    "tarifa_anterior":          "120",
    "precio_bencina_anterior":  "1188",
    "delta_frecuencia":         "2",
}

# ─────────────────────────────────────────
# CREAR TABLAS
# ─────────────────────────────────────────

def crear_tablas(c):
    c.executescript("""
    CREATE TABLE IF NOT EXISTS auditor (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre        TEXT    NOT NULL UNIQUE,
        ciudad        TEXT,
        km_por_litro  REAL    NOT NULL,
        activo        INTEGER DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS ruta (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        numero       INTEGER,
        nombre       TEXT    NOT NULL,
        distancia_km INTEGER NOT NULL,
        id_auditor   INTEGER NOT NULL,
        FOREIGN KEY (id_auditor) REFERENCES auditor(id)
    );

    CREATE TABLE IF NOT EXISTS historial_km (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        id_auditor   INTEGER NOT NULL,
        anio         INTEGER NOT NULL,
        mes          INTEGER NOT NULL,
        km_realizados REAL   NOT NULL,
        UNIQUE(id_auditor, anio, mes),
        FOREIGN KEY (id_auditor) REFERENCES auditor(id)
    );

    CREATE TABLE IF NOT EXISTS presupuesto_anual (
        id                 INTEGER PRIMARY KEY AUTOINCREMENT,
        anio               INTEGER NOT NULL,
        mes                INTEGER NOT NULL,
        presupuesto_total  REAL    NOT NULL,
        UNIQUE(anio, mes)
    );

    CREATE TABLE IF NOT EXISTS parametro_mes (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        anio            INTEGER NOT NULL,
        mes             INTEGER NOT NULL,
        precio_bencina  REAL,
        dias_habiles    INTEGER,
        UNIQUE(anio, mes)
    );

    CREATE TABLE IF NOT EXISTS programacion_mensual (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        id_auditor INTEGER NOT NULL,
        anio       INTEGER NOT NULL,
        mes        INTEGER NOT NULL,
        id_ruta    INTEGER NOT NULL,
        UNIQUE(id_auditor, anio, mes, id_ruta),
        FOREIGN KEY (id_auditor) REFERENCES auditor(id),
        FOREIGN KEY (id_ruta)    REFERENCES ruta(id)
    );

    CREATE TABLE IF NOT EXISTS registro_pago (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        id_auditor INTEGER NOT NULL,
        anio       INTEGER NOT NULL,
        mes        INTEGER NOT NULL,
        bono       REAL,
        pago_km    REAL,
        km_con_bono REAL,
        UNIQUE(id_auditor, anio, mes),
        FOREIGN KEY (id_auditor) REFERENCES auditor(id)
    );

    CREATE TABLE IF NOT EXISTS config (
        clave TEXT PRIMARY KEY,
        valor TEXT
    );
    """)

# ─────────────────────────────────────────
# CARGA DE DATOS
# ─────────────────────────────────────────

def cargar_auditores(c):
    for nombre, d in AUDITORES_ACTIVOS.items():
        c.execute("""INSERT OR IGNORE INTO auditor (nombre, ciudad, km_por_litro, activo)
                     VALUES (?, ?, ?, 1)""",
                  (nombre, d["ciudad"], d["rendimiento"]))
    for nombre, d in AUDITORES_INACTIVOS.items():
        c.execute("""INSERT OR IGNORE INTO auditor (nombre, ciudad, km_por_litro, activo)
                     VALUES (?, ?, ?, 0)""",
                  (nombre, d["ciudad"], d["rendimiento"]))
    print(f"  Auditores cargados: {len(AUDITORES_ACTIVOS)} activos, {len(AUDITORES_INACTIVOS)} inactivos")


def cargar_rutas(c):
    wb = load_workbook(EXCEL_RUTAS, read_only=True)
    total = 0
    for sheet in wb.sheetnames:
        row_aud = c.execute("SELECT id FROM auditor WHERE nombre=?", (sheet,)).fetchone()
        if not row_aud:
            print(f"  ⚠️  Hoja '{sheet}' no coincide con ningún auditor, se omite")
            continue
        id_auditor = row_aud[0]
        ws = wb[sheet]
        for fila in ws.iter_rows(min_row=4, values_only=True):
            vals = [v for v in fila if v is not None]
            if len(vals) >= 3 and isinstance(vals[0], (int, float)) and isinstance(vals[-1], (int, float)):
                num, nombre_ruta, km = vals[0], vals[1], vals[-1]
                if isinstance(num, int) and 1 <= km <= 500:
                    c.execute("""INSERT OR IGNORE INTO ruta (numero, nombre, distancia_km, id_auditor)
                                 VALUES (?, ?, ?, ?)""",
                              (int(num), str(nombre_ruta).strip(), int(km), id_auditor))
                    total += 1
    wb.close()
    print(f"  Rutas cargadas: {total}")


def cargar_historico(c):
    wb = load_workbook(EXCEL_HISTORICO, read_only=True)
    ws = wb["Resumen mensual"]
    total = 0
    for fila in ws.iter_rows(min_row=5, values_only=True):
        vals = list(fila)
        mes_val = next((v for v in vals if isinstance(v, str) and v in MESES_NUM), None)
        if not mes_val:
            continue
        fila_str = " ".join(str(v) for v in vals if v is not None)
        if "Km" not in fila_str or "2025" not in fila_str:
            continue
        nums = [v for v in vals if isinstance(v, (int, float)) and v > 0]
        if len(nums) < 6:
            continue
        for i, nombre in enumerate(AUDITORES_ORDEN_HIST):
            row_aud = c.execute("SELECT id FROM auditor WHERE nombre=?", (nombre,)).fetchone()
            if not row_aud:
                continue
            c.execute("""INSERT OR IGNORE INTO historial_km (id_auditor, anio, mes, km_realizados)
                         VALUES (?, ?, ?, ?)""",
                      (row_aud[0], 2025, MESES_NUM[mes_val], nums[i]))
            total += 1
    wb.close()
    print(f"  Histórico km cargado: {total} registros")


def cargar_presupuesto(c):
    total = 0
    for anio, meses in PRESUPUESTO.items():
        for mes, monto in meses.items():
            c.execute("""INSERT OR IGNORE INTO presupuesto_anual (anio, mes, presupuesto_total)
                         VALUES (?, ?, ?)""", (anio, mes, monto))
            total += 1
    print(f"  Presupuesto cargado: {total} registros")


def cargar_config(c):
    for clave, valor in CONFIG.items():
        c.execute("INSERT OR IGNORE INTO config VALUES (?, ?)", (clave, valor))
    print(f"  Config cargada: {len(CONFIG)} parámetros")


# ─────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────

def main():
    print("=== INICIALIZANDO BASE DE DATOS ===\n")

    # Verificar que los Excel existen
    for path, nombre in [(EXCEL_RUTAS, "Rutas"), (EXCEL_HISTORICO, "Histórico")]:
        if not os.path.exists(path):
            print(f"❌ No se encontró el archivo: {path}")
            print(f"   Asegúrate de poner el Excel de {nombre} en la carpeta app\\")
            return

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    crear_tablas(c)
    print("Tablas creadas.\n")
    print("Cargando datos:")
    cargar_auditores(c)
    cargar_rutas(c)
    cargar_historico(c)
    cargar_presupuesto(c)
    cargar_config(c)
    conn.commit()

    # Resumen
    print("\n=== RESUMEN FINAL ===")
    for tabla in ["auditor", "ruta", "historial_km", "presupuesto_anual", "config"]:
        n = c.execute(f"SELECT COUNT(*) FROM {tabla}").fetchone()[0]
        print(f"  {tabla}: {n} registros")

    conn.close()
    print(f"\n✅ Base de datos lista en: {DB_PATH}")
    print("   Copia roble.db a tu carpeta de Google Drive para compartirla.")


if __name__ == "__main__":
    main()